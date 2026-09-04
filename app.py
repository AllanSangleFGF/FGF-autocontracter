from flask import Flask, redirect, send_from_directory, url_for, render_template, request, jsonify, send_file, abort, session, make_response
from flask_session import Session
from playwright.sync_api import sync_playwright
from functools import wraps
import openpyxl
import os

app = Flask(__name__)
app.secret_key = "secret123"
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

phone_loc  = "phone_contracts"
laptop_loc = "laptop_contracts"
EXCEL_PATH        = "phone_list.xlsx"          # ← phone roster
LAPTOP_EXCEL_PATH = "laptop_list.xlsx"  # ← laptop roster


# ── LOGIN REQUIRED DECORATOR ─────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("username"):
            return redirect(url_for('home'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("username"):
            return redirect(url_for('home'))
        if session.get("username") != "admin":
            abort(403)
        return f(*args, **kwargs)
    return decorated


# ── HELPERS ──────────────────────────────────────────────────────────────────
def load_users(filepath="users.txt"):
    users = {}
    if not os.path.exists(filepath):
        return users
    with open(filepath, "r") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split(":", 2)
            if len(parts) == 3:
                uname, pwd, display = parts
                users[uname.strip().lower()] = {
                    "password": pwd.strip(),
                    "display_name": display.strip()
                }
    return users


def load_team_members(filepath=EXCEL_PATH):
    """Read all rows from the phone Excel roster into a list of dicts."""
    members = []
    if not os.path.exists(filepath):
        return members
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        member = {h: (str(v).strip() if v is not None else "") for h, v in zip(headers, row)}
        members.append(member)
    return members


def load_laptop_members(filepath=LAPTOP_EXCEL_PATH):
    """Read all rows from the laptop Excel roster into a list of dicts."""
    members = []
    if not os.path.exists(filepath):
        return members
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        member = {h: (str(v).strip() if v is not None else "") for h, v in zip(headers, row) if h is not None}
        members.append(member)
    return members


# ── PUBLIC ROUTES ─────────────────────────────────────────────────────────────
@app.route('/')
def home():
    if session.get("username"):
        return redirect(url_for('main_dashboard'))
    return render_template('login.html', the_title='FGF TechCare Login')


@app.route('/login.html', methods=['POST'])
def login():
    username = request.form.get("username", "").strip().lower()
    password = request.form.get("password", "").strip()
    print("LOGIN ATTEMPT:", username, password)

    users = load_users()
    user  = users.get(username)

    if username == "admin" and password == '1234':
        session["username"]     = username
        session["display_name"] = user["display_name"]
        response = make_response(redirect(url_for('admin_contracts')))
        response.set_cookie("Admin", user["display_name"], samesite="Lax")
        print("Admin login successful")
        return response

    if user and user["password"] == password:
        session["username"]     = username
        session["display_name"] = user["display_name"]
        response = make_response(redirect(url_for('main_dashboard')))
        response.set_cookie("agent_name", user["display_name"], samesite="Lax")
        print("Regular login successful")
        return response
    else:
        return render_template('login.html', the_title='FGF TechCare Login', error="Invalid username or password.")


@app.route('/logout')
def logout():
    session.clear()
    response = make_response(redirect(url_for('home')))
    response.delete_cookie("agent_name")
    return response


# ── PROTECTED ROUTES ──────────────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def main_dashboard():
    return render_template('dashboard.html', the_title='Dashboard',
                           username=session.get("username"),
                           display_name=session.get("display_name", ""))


@app.route('/phone-contract')
@login_required
def phone_contract():
    return render_template('phone_contract.html', the_title='Phone Contract',
                           username=session.get("username"),
                           display_name=session.get("display_name", ""))


@app.route('/laptop-contract')
@login_required
def laptop_contracts():
    return render_template('laptop_contract.html', the_title='Laptop Use Policy')

@app.route('/multi-contract')
@login_required
def multi_contracts():
    return render_template('multi_contract.html', the_title='Multi Contract')

# ── ADMIN ROUTES ──────────────────────────────────────────────────────────────
@app.route('/admin/upload-roster', methods=['GET'])
@admin_required
def upload_roster_page():
    return render_template('admin_upload.html', the_title='Upload Roster')


@app.route('/admin/upload-roster', methods=['POST'])
@admin_required
def upload_roster():
    file_type = request.form.get('type', '')   # 'phone' or 'laptop'
    f = request.files.get('file')
    if not f or not f.filename.endswith('.xlsx'):
        return jsonify({"status": "error", "message": "Please upload a valid .xlsx file"}), 400
    dest = EXCEL_PATH if file_type == 'phone' else LAPTOP_EXCEL_PATH
    f.save(dest)
    return jsonify({"status": "success", "message": f"{'Phone' if file_type == 'phone' else 'Laptop'} roster updated successfully."})


@app.route('/admin/contracts')
@admin_required
def admin_contracts():
    import datetime
    results = []
    for folder, label in [(phone_loc, 'Phone'), (laptop_loc, 'Laptop')]:
        if os.path.exists(folder):
            for fname in sorted(os.listdir(folder)):
                if fname.endswith('.pdf'):
                    fpath = os.path.join(folder, fname)
                    mtime = os.path.getmtime(fpath)
                    modified = datetime.datetime.fromtimestamp(mtime).strftime('%b %d, %Y')
                    results.append({
                        "name":     fname,
                        "type":     label,
                        "path":     fpath.replace("\\", "/"),
                        "size":     round(os.path.getsize(fpath) / 1024, 1),
                        "modified": modified,
                    })
    results.sort(key=lambda r: r['name'].lower())
    return render_template('admin_contracts.html', the_title='Search Contracts', results=results)

@app.route("/downloadp")
def download_file_phone():
    FILE_DIRECTORY = "files"
    FILE_NAME = "base_phone_template.xlsx"
    
    return send_from_directory(
        FILE_DIRECTORY,
        FILE_NAME,
        as_attachment=True
    )

@app.route("/downloadl")
def download_file_laptop():
    FILE_DIRECTORY = "files"

    FILE_NAME = "base_laptop_template.xlsx"

    return send_from_directory(
        FILE_DIRECTORY,
        FILE_NAME,
        as_attachment=True
    )



# ── EXCEL SEARCH API ──────────────────────────────────────────────────────────
@app.route('/api/search-member')
@login_required
def search_member():
    q     = request.args.get("q",     "").strip().lower()
    exact = request.args.get("exact", "").strip().lower()
    members = load_team_members()

    if exact:
        for m in members:
            if m.get("Team_Member_Name", "").lower() == exact or \
               m.get("Team_Member_Usern", "").lower() == exact:
                return jsonify({"status": "found", "member": m})
        return jsonify({"status": "not_found"}), 404

    if q:
        results = []
        for m in members:
            name  = m.get("Team_Member_Name", "")
            uname = m.get("Team_Member_Usern", "")
            if q in name.lower() or q in uname.lower():
                results.append({"display": f"{name} ({uname})", "name": name, "username": uname})
        return jsonify({"status": "ok", "results": results})

    return jsonify({"status": "ok", "results": []})


# ── LAPTOP EXCEL SEARCH API ───────────────────────────────────────────────────
@app.route('/api/search-laptop-member')
@login_required
def search_laptop_member():
    q     = request.args.get("q",     "").strip().lower()
    exact = request.args.get("exact", "").strip().lower()
    members = load_laptop_members()

    if exact:
        for m in members:
            if m.get("Team_Member_Name", "").lower() == exact or \
               m.get("Team_Member_Usern", "").lower() == exact:
                return jsonify({"status": "found", "member": m})
        return jsonify({"status": "not_found"}), 404

    if q:
        results = []
        for m in members:
            name  = m.get("Team_Member_Name", "")
            uname = m.get("Team_Member_Usern", "")
            if q in name.lower() or q in uname.lower():
                results.append({"display": f"{name} ({uname})", "name": name, "username": uname})
        return jsonify({"status": "ok", "results": results})

    return jsonify({"status": "ok", "results": []})


# ── PDF SUBMIT ROUTES ─────────────────────────────────────────────────────────
@app.route("/phone-submit", methods=["POST"])
@login_required
def phone_submit():
    data        = request.json
    agent_name  = data.get("name", "Contract")
    phone_model = data.get("phone_model", "Phone")

    os.makedirs(phone_loc, exist_ok=True)
    pdf_path     = f"{phone_loc}/{agent_name} - {phone_model}.pdf"
    html_content = data.get("html")
    if not html_content:
        return jsonify({"status": "error", "message": "No HTML content provided"}), 400

    with sync_playwright() as p:
        browser = p.chromium.launch()
        page    = browser.new_page()
        page.set_content(html_content, wait_until="networkidle")
        page.pdf(path=pdf_path, format="A4", print_background=True,
                 margin={"top": "0", "right": "0", "bottom": "0", "left": "0"})
        browser.close()

    return jsonify({"status": "success", "file": pdf_path})


@app.route("/laptop-submit", methods=["POST"])
@login_required
def laptop_submit():
    data         = request.json
    agent_name   = data.get("name", "Contract")
    laptop_model = data.get("laptop_name", "laptop")

    os.makedirs(laptop_loc, exist_ok=True)
    pdf_path     = f"{laptop_loc}/{agent_name} - {laptop_model}.pdf"
    html_content = data.get("html")
    if not html_content:
        return jsonify({"status": "error", "message": "No HTML content provided"}), 400

    with sync_playwright() as p:
        browser = p.chromium.launch()
        page    = browser.new_page()
        page.set_content(html_content, wait_until="networkidle")
        page.pdf(path=pdf_path, format="A4", print_background=True,
                 margin={"top": "0", "right": "0", "bottom": "0", "left": "0"})
        browser.close()

    return jsonify({"status": "success", "file": pdf_path})


# ── DOWNLOAD ─────────────────────────────────────────────────────────────────
@app.route("/download")
@login_required
def download():
    file_path = request.args.get("file", "")
    abs_path  = os.path.abspath(file_path)

    # Allow downloads from either contracts folder
    allowed = [os.path.abspath(phone_loc), os.path.abspath(laptop_loc)]
    if not any(abs_path.lower().startswith(d.lower()) for d in allowed):
        abort(403)
    if not os.path.exists(abs_path):
        abort(404)

    return send_file(abs_path, as_attachment=True,
                     download_name=os.path.basename(abs_path),
                     mimetype="application/pdf")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=80, debug=True)